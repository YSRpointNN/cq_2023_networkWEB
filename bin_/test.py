# import openpyxl

# def server_find_from_excl(ws_name:str, col_i:int, find_val:str):
#     sourceWB = openpyxl.load_workbook('E:/MyCode/networkWeb/device_list.xlsx')
#     sourceWS = sourceWB[ws_name]
#     retrun_info = list()
#     for row in sourceWS.iter_rows(min_row=2, values_only=True):
#         if row[col_i] == find_val:
#             if retrun_info:
#                 return '多個匹配數據'
#             else:
#                 retrun_info = row
#     if retrun_info:
#         return retrun_info
#     else:
#         return '未找到'

# a = server_find_from_excl('DL2核心及Server', 0, '23080009JDV')
# print(a)

from openpyxl import load_workbook

workbook = load_workbook('device_list.xlsx')
sheet_list = workbook.sheetnames

for sheet_name in sheet_list:
    sheet = workbook[sheet_name]
    for row in sheet.iter_rows(values_only=True):
        if 'BN01' in str(row):
            print(sheet_name)
            print(row)

print('END-')
