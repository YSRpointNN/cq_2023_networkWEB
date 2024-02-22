import openpyxl, datetime

# 监控清单_读sheet数据 返回 [直接字典,..]
def devl_ret(sheet_name = '') -> tuple:
    sourceWB = openpyxl.load_workbook(r'./ExcelBase/device_list.xlsx', read_only=True)
    redata = list()
    sheet_list = sourceWB.sheetnames
    if sheet_name:
        pass
    else:
        sheet_name = sheet_list[0]
    sourceWS = sourceWB[sheet_name]
    i = 1
    for row in sourceWS.iter_rows(min_row=2, values_only=True):
        redata.append({'index':i, 'sheet':sheet_name, 'devi':row[0], 'loca1':row[1], 'loca2':row[2], 'model':row[3], 'name':row[4], 'ip':row[5], 'assets':row[6], 'sn':row[7], 'mark':row[8]})
        i += 1
    sourceWB.close()
    return redata, sheet_list

# 固资异动_读sheet数据 返回 [直接字典,..]
def devm_ret(date) -> list:
    sourceWB = openpyxl.load_workbook(r'./ExcelBase/device_move.xlsx', read_only=True, data_only=True)
    redata = list()
    sheet_list = sourceWB.sheetnames

    sheet_name = False
    for i in sheet_list:
        if i == date[:4]:
            sheet_name  = i
    if not sheet_name:
        return []
    sourceWS = sourceWB[sheet_name]
    for row in sourceWS.iter_rows(min_row=2, values_only=True):
        debug = row[10]
        if type(debug) == datetime.datetime:
            actdate = row[10].strftime("%Y-%m-%d")
            if actdate[:7] == date:
                redata.append({'devi':row[0], 'timebar':actdate,'type':row[3], 'model':row[4], 'owner':row[5], 'name':row[6], 'sn':row[2], 'ip':row[7], 'assets':row[1], 'locas':row[8], 'locan':row[9], 'opter':row[11], 'mark':row[12]})
        elif type(debug) == str:
            actdate = row[10]
            if actdate[:7] == date:
                redata.append({'devi':row[0], 'timebar':actdate,'type':row[3], 'model':row[4], 'owner':row[5], 'name':row[6], 'sn':row[2], 'ip':row[7], 'assets':row[1], 'locas':row[8], 'locan':row[9], 'opter':row[11], 'mark':row[12]})
        else:
            pass
    sourceWB.close()
    return redata

# 库存_读sheet数据 返回 [直接字典,..]
def deva_ret() -> list:
    sourceWB = openpyxl.load_workbook(r'./ExcelBase/device_all.xlsx', read_only=True)
    redata = list()
    sheet_list = sourceWB.sheetnames
    sheet_name = sheet_list[0]
    sourceWS = sourceWB[sheet_name]
    i = 0
    for row in sourceWS.iter_rows(min_row=1, values_only=True):
        redata.append({'index':i, 'devi':row[0], 'assets':row[1], 'sn':row[2], 'loca2':row[3], 'type':row[4], 'model':row[5], 'specs':row[6], 'net':row[7], 'owner':row[8], 'hook':row[9], 'cunfang':row[10],})
        i += 1
    sourceWB.close()
    return redata

# 传入uid字符串  返回 (行数,直接字典)
def deva_uidfind(uid:str) -> tuple:
    sourceWB = openpyxl.load_workbook(r'./ExcelBase/device_all.xlsx', read_only=True)
    rbline = ''
    rbindex = 0
    for activ_name in sourceWB.sheetnames:
        activWS = sourceWB[activ_name]
        rbindex = 0
        for line in activWS.iter_rows(min_row=0, values_only=True):
            rbindex += 1
            if line[0] == uid:
                rbline = {'assets':line[1], 'sn':line[2], 'loca2':line[3], 'type':line[4], 'model':line[5], 'owner':line[8]}
                break
    sourceWB.close()
    return 0, rbline

# 传入固资字符串  返回 (行数,直接字典)
def deva_assetfind(asset:str) -> tuple:
    sourceWB = openpyxl.load_workbook(r'./ExcelBase/device_all.xlsx', read_only=True)
    rbindex = 0
    for activ_name in sourceWB.sheetnames:
        activWS = sourceWB[activ_name]
        for line in activWS.iter_rows(min_row=1, values_only=True):
            rbindex += 1
            if line[1] == asset:
                rbline = {'devi':line[0],'assets':line[1], 'sn':line[2], 'loca2':line[3], 'type':line[4], 'model':line[5], 'owner':line[8]}
                sourceWB.close()
                return rbindex, rbline
    return False, []

# 传入uid字符串  返回 (行数,工作表名称)
def devl_uidfind(find_uid:str) -> tuple:
    sourceWB = openpyxl.load_workbook(r'./ExcelBase/device_list.xlsx', read_only=True)
    findrb = None
    for activ_name in sourceWB.sheetnames:
        activWS = sourceWB[activ_name]
        for line in activWS.iter_rows(min_row=0):
            if line[0].value == find_uid:
                findrb = line[0].row
                break
    sourceWB.close()
    return findrb, activ_name

# 用户查找 返回 [直接字典,..]
def devl_usrfind(find_usr:str) -> list:
    if find_usr:
        sourceWB = openpyxl.load_workbook(r'./ExcelBase/device_list.xlsx', read_only=True)
        sheet_list = sourceWB.sheetnames
        find_s_load = list()
        for sheet_name in sheet_list:
            sheet = sourceWB[sheet_name]
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if find_usr in str(row):
                    find_s_load.append({'index':i, 'sheet':sheet_name, 'devi':row[0], 'loca1':row[1], 'loca2':row[2], 'model':row[3], 'name':row[4], 'ip':row[5], 'assets':row[6], 'sn':row[7], 'mark':row[8]})
        sourceWB.close()
        return find_s_load
    return None

# 用户查找 返回 [直接字典,..]
def deva_usrfind(find_usr:str) -> list:
    if find_usr:
        sourceWB = openpyxl.load_workbook(r'./ExcelBase/device_all.xlsx', read_only=True)
        find_s_load = list()
        sheet = sourceWB['deva']
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if find_usr in str(row):
                find_s_load.append({'devi':row[0], 'assets':row[1], 'sn':row[2], 'loca2':row[3], 'type':row[4], 'model':row[5]})
        sourceWB.close()
        return find_s_load
    else:
        return None


# 通用uid查找 返回 (行数, 工作表名称)
def GEN_uidfind(wb:str, find_uid:str) -> tuple:
    sourceWB = openpyxl.load_workbook(wb, read_only=True)
    findrb = None
    sheet = None
    for activ_name in sourceWB.sheetnames:
        activWS = sourceWB[activ_name]
        for line in activWS.iter_rows(min_row=0):
            if line[0].value == find_uid:
                findrb = line[0].row
                sheet = activ_name
                break
    sourceWB.close()
    return findrb, sheet

# 通用编辑 返回 True/False
def GEN_edit(wbpath, line) -> bool:
    devi = line[0]
    edWB = openpyxl.load_workbook(wbpath, read_only=False)
    try:
        index, sheet = GEN_uidfind(wbpath, devi)
        edWS = edWB[sheet]
        for i, value in enumerate(line, start=1):
            cell = edWS.cell(row=index, column=i)
            cell.value = value
        edWB.save(wbpath)
        edWB.close()
    except:
        return False
        edWB.close()
    return True

# 通用删除
def GEN_delete(wbpath, devis) -> bool:
    delWB = openpyxl.load_workbook(wbpath, read_only=False)
    for devi in devis:
        index, sheet = GEN_uidfind(wbpath, devi)
        delWS = delWB[sheet]
        delWS.delete_rows(index)
    delWB.save(wbpath)
    delWB.close()

def devl_append(data) -> bool:
    apWB = openpyxl.load_workbook(r'./ExcelBase/device_list.xlsx', read_only=False)
    for l in data:
        sheet_name = l['sheet']
        sheet_index = int(l['index']) + 1
        linelist = (l['devi'], l['loca1'], l['loca2'], l['model'], l['name'], l['ip'], l['assets'], l['sn'], l['mark'])
        actWS = apWB[sheet_name]

        actWS.insert_rows(sheet_index, 1)
        for i in range(len(linelist)):
            actWS.cell(row=sheet_index, column=i+1).value = linelist[i]
    apWB.save(r'./ExcelBase/device_list.xlsx')
    apWB.close()
    return True

def devm_append(data) -> bool:
    apWB = openpyxl.load_workbook(r'./ExcelBase/device_move.xlsx', read_only=False)
    for index in data:
        l = data[index]
        seddate = l['sedmvdate']
        if len(seddate) < 2:
            return False
        
        sheet_name = seddate[:4]

        toexc_date = datetime.datetime.strptime(seddate, '%Y-%m-%d')
        try:
            actWS = apWB[sheet_name]
        except KeyError:
            apWB.create_sheet(sheet_name)
            actWS = apWB[sheet_name]
        except:
            return False
        sheet_index = actWS.max_row + 1
        linelist = (index[4:], l['assets'], l['sn'], l['type'], l['model'], l['sedmvowner'], l['name'], l['ip'], l['loca2'], l['sedmvloca'], toexc_date, l['sedmvopter'], l['sedmvmark'])

        actWS.insert_rows(sheet_index, 1)
        for i in range(len(linelist)):
            actWS.cell(row=sheet_index, column=i+1).value = linelist[i]
    apWB.close()
    apWB.save(r'./ExcelBase/device_move.xlsx')
    return True



def GEN_newsheet(path, head):
    pass

def devm_script_cindex():
    pass
